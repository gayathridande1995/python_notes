from openpyxl import load_workbook
wb2 = load_workbook('sample.xlsx')
ws=wb2.get_sheet_by_name('Sheet')
print(ws['A1'].value)

for rowOfCellObjects in ws['A1':'C3']:
    for cellObj in rowOfCellObjects:
               print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')


ws['A1'].value=123

wb2.save('sample.xlsx')